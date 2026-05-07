#!/usr/bin/env python3
"""
상세 기능명세서 엑셀 생성 스크립트 v2

사용법:
    python create_spec_xlsx.py <project_name> <output_path>

이 스크립트는 빈 템플릿(인우님 Quiket v1.0 명세서 구조)을 생성한다.
실제 데이터는 SKILL.md의 절차에 따라 AI가 채운다.

v2 변경사항:
- Sheet 5 "기능명세 질의" 추가 (개발자·디자이너 ↔ 기획자 Q&A 채널)
- DB구조, 자격증 종류 도메인 데이터셋 자리 추가
- 상세 시트 헤더 행 1에 항목 가이드 추가
- 컬럼 폭 인우님 v1.0 기준으로 재조정
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="E8E8E8")
HEADER_FONT = Font(name="Arial", size=11, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True)
GUIDE_FONT = Font(name="Arial", size=9, italic=True)
DEFAULT_FONT_NAME = "Arial"


def style_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_spec_template(project_name: str, output_path: str, version: str = "1.0"):
    wb = Workbook()
    wb.remove(wb.active)

    # ==========================================================
    # Sheet 1: 기능명세_목록
    # ==========================================================
    ws1 = wb.create_sheet(f"기능명세_목록 v.{version}")

    ws1.cell(row=1, column=1, value=f"{project_name} — 기능 명세 목록  |  v{version}")
    ws1.cell(row=1, column=1).font = TITLE_FONT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    headers_list = ["ID", "섹션", "기능명", "한줄 설명", "포함\n여부", "우선\n순위", "관련 US"]
    for col_idx, header in enumerate(headers_list, start=1):
        style_header_cell(ws1.cell(row=2, column=col_idx, value=header))

    col_widths_list = {1: 18, 2: 14, 3: 26, 4: 55, 5: 10, 6: 10, 7: 14}
    for col, width in col_widths_list.items():
        ws1.column_dimensions[get_column_letter(col)].width = width

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 32
    ws1.freeze_panes = "A3"

    # ==========================================================
    # Sheet 2: 기능명세_상세
    # ==========================================================
    ws2 = wb.create_sheet(f"기능명세_상세 v.{version}")

    ws2.cell(row=1, column=1, value=f"{project_name} — 기능 명세 상세  | v.{version}")
    ws2.cell(row=1, column=1).font = TITLE_FONT
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    guide_text = (
        "[항목]\n"
        "- 설명/입력/출력/정책/예외/플로우\n"
        "- #: 케이스별 분기 (예: #1 토큰유효, #객관식, #한문제씩 모드)\n"
        "- (Step1), (A-Step2): 다단계 입력 분기\n"
        "- 입력: 앱 → 서버 / 출력: 서버 → 앱"
    )
    ws2.cell(row=1, column=6, value=guide_text)
    ws2.cell(row=1, column=6).font = GUIDE_FONT
    ws2.cell(row=1, column=6).alignment = Alignment(wrap_text=True, vertical="center")
    ws2.merge_cells(start_row=1, start_column=6, end_row=1, end_column=10)

    headers_detail = [
        "기능ID\n(#화면ID)", "섹션", "기능명", "항목", "내용",
        "우선순위", "API 연동", "포함여부", "상태", "메모/이슈"
    ]
    for col_idx, header in enumerate(headers_detail, start=1):
        style_header_cell(ws2.cell(row=2, column=col_idx, value=header))

    col_widths_detail = {1: 22, 2: 14, 3: 24, 4: 18, 5: 75, 6: 10, 7: 22, 8: 10, 9: 12, 10: 38}
    for col, width in col_widths_detail.items():
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws2.row_dimensions[1].height = 70
    ws2.row_dimensions[2].height = 32
    ws2.freeze_panes = "A3"

    # ==========================================================
    # Sheet 3: DB구조 (자리만 잡아두기)
    # ==========================================================
    ws_db = wb.create_sheet("DB구조")
    ws_db.cell(row=1, column=1, value=f"{project_name} — DB 구조 (별도 작성)")
    ws_db.cell(row=1, column=1).font = TITLE_FONT
    ws_db.column_dimensions["A"].width = 60

    # ==========================================================
    # Sheet 4: 자격증 종류 (도메인 데이터셋 자리)
    # ==========================================================
    ws_cert = wb.create_sheet("자격증 종류")
    ws_cert.cell(row=1, column=1, value=f"{project_name} — 도메인 데이터셋 (자격증 종류 등, 별도 작성)")
    ws_cert.cell(row=1, column=1).font = TITLE_FONT
    ws_cert.column_dimensions["A"].width = 60

    # ==========================================================
    # Sheet 5: 기능명세 질의 (v2 신규) ★
    # ==========================================================
    ws_q = wb.create_sheet("기능명세 질의")

    headers_q = ["순번", "기능ID / (#화면ID)", "섹션", "기능명", "문의 제목", "문의사항", "답변"]
    for col_idx, header in enumerate(headers_q, start=1):
        style_header_cell(ws_q.cell(row=1, column=col_idx, value=header))

    col_widths_q = {1: 8, 2: 22, 3: 14, 4: 22, 5: 28, 6: 60, 7: 60}
    for col, width in col_widths_q.items():
        ws_q.column_dimensions[get_column_letter(col)].width = width

    ws_q.row_dimensions[1].height = 30

    # 순번 1~30 미리 채우기
    for i in range(1, 31):
        ws_q.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")

    ws_q.freeze_panes = "B2"

    # ==========================================================
    # Sheet 6: 변경 이력
    # ==========================================================
    ws_hist = wb.create_sheet("변경 이력")
    headers_hist = ["버전", "일자", "변경 내용", "변경자"]
    for col_idx, header in enumerate(headers_hist, start=1):
        style_header_cell(ws_hist.cell(row=1, column=col_idx, value=header))

    col_widths_hist = {1: 10, 2: 14, 3: 60, 4: 14}
    for col, width in col_widths_hist.items():
        ws_hist.column_dimensions[get_column_letter(col)].width = width

    ws_hist.row_dimensions[1].height = 28

    # 초기 행 (현재 버전 기록 자리)
    ws_hist.cell(row=2, column=1, value=f"v{version}")
    ws_hist.cell(row=2, column=3, value="최초 작성")

    # ==========================================================
    # 저장
    # ==========================================================
    wb.save(output_path)
    print(f"✅ 명세서 템플릿 생성 완료: {output_path}")
    print(f"   - Sheet 1: 기능명세_목록 v.{version}")
    print(f"   - Sheet 2: 기능명세_상세 v.{version}")
    print(f"   - Sheet 3: DB구조 (자리)")
    print(f"   - Sheet 4: 자격증 종류 (자리)")
    print(f"   - Sheet 5: 기능명세 질의 (순번 1~30 미리 채움)  ★ v2 신규")
    print(f"   - Sheet 6: 변경 이력")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python create_spec_xlsx.py <project_name> <output_path> [version]")
        sys.exit(1)

    project_name_arg = sys.argv[1]
    output_path_arg = sys.argv[2]
    version_arg = sys.argv[3] if len(sys.argv) >= 4 else "1.0"
    create_spec_template(project_name_arg, output_path_arg, version_arg)
