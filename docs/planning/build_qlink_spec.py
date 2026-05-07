#!/usr/bin/env python3
"""
QLINK 상세 기능명세서 v1.0 — xlsx 빌더

inwoo-spec-writer 패턴(Quiket v1.0 레퍼런스)에 따라 6개 시트 생성:
  1. 기능명세_목록 v.1.0
  2. 기능명세_상세 v.1.0
  3. DB구조           — 별도 작성 자리
  4. 도메인 데이터셋    — 별도 작성 자리
  5. 기능명세 질의     — 1~30 순번 미리 채움
  6. 변경 이력

데이터는 본 스크립트 인라인. 05_상세기능명세서.md와 일치.

사용법:
  cd docs/planning && python3 build_qlink_spec.py
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_NAME = "QLINK"
VERSION = "1.0"
TODAY = "2026-05-07"

# ===== styles =====
HEADER_FILL = PatternFill("solid", start_color="E8E8E8")
HEADER_FONT = Font(name="Arial", size=11, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True)
GUIDE_FONT = Font(name="Arial", size=9, italic=True)
GRAY_FONT = Font(color="888888")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = WRAP_CENTER

# ===== Sheet 1: 기능명세_목록 데이터 =====
LIST_ROWS = [
    # ID, 섹션, 기능명, 한줄 설명, 포함여부, 우선순위, 관련 US
    ("AUTH-001", "인증", "이메일 회원가입", "이메일·비밀번호·닉네임으로 계정 생성, 인증 메일 검증 포함", "MVP", "Must", "US-001"),
    ("AUTH-002", "인증", "이메일 로그인", "이메일·비밀번호로 로그인", "MVP", "Must", "US-001"),
    ("AUTH-003", "인증", "게스트 모드 (프로토타입 전용)", "프로토타입 검토용. 출시 시 제거", "프로토타입", "Won't", "US-011"),
    ("AUTH-004", "인증", "자동 로그인 / 토큰 갱신", "리프레시 토큰으로 무중단 세션", "MVP", "Must", "US-001"),
    ("AUTH-005", "인증", "소셜 로그인 (Google/Apple/Kakao)", "OAuth 가입/로그인 — MVP 보류", "보류", "Should", "US-001"),
    ("LINK-001", "링크", "URL 직접 입력 추가", "링크 추가 시트에서 URL 저장", "MVP", "Must", "US-001"),
    ("LINK-002", "링크", "클립보드 자동 감지", "시트 진입 시 클립보드 URL 자동 채움", "MVP", "Must", "US-001"),
    ("LINK-003", "링크", "시스템 공유 시트 수신", "Web Share Target 처리", "MVP", "Must", "US-001"),
    ("LINK-004", "링크", "AI 한 줄 요약 + 자동 태그", "저장 직후 AI 요약·태그 생성", "MVP", "Must", "US-002"),
    ("LINK-005", "링크", "링크 상세 (열기·공유·이동·삭제)", "카드 탭 → 시트, 5가지 액션", "MVP", "Must", "US-003"),
    ("LINK-006", "링크", "링크 메모/제목 수정", "제목·요약·태그 수동 편집", "MVP", "Must", "US-003"),
    ("LINK-007", "링크", "(DEPRECATED) 단일 알림 예약", "TODO-001~005에 통합", "미포함", "Won't", ""),
    ("TODO-001", "할일", "링크당 다중 할 일", "한 링크에 N개(≤20). 제목+3가지 알림 모드(없음/시간선택/반복)", "MVP", "Must", "US-007, US-009"),
    ("TODO-002", "할일", "반복 알림 (요일 기반)", "일~토 체크박스(default 매일) + 시간 + 종료일", "MVP", "Must", "US-008"),
    ("TODO-003", "할일", "할 일 완료 체크 (회차별 + 누적)", "반복은 회차별 완료, 누적 이력 토글로 접기·펼치기", "MVP", "Must", "US-011"),
    ("TODO-004", "할일", "할일 통합 화면", "모든 할 일 평탄화. 필터 5종", "MVP", "Must", "US-010"),
    ("TODO-005", "할일", "할 일 알림 발송 (FCM)", "EventBridge → Lambda → FCM Topic. 반복 자동 스케줄", "MVP", "Must", "US-008"),
    ("TODO-009", "할일", "할 일 가시성 (공유 폴더)", "🔒나만 / 👥공유 선택. 메모는 항상 비공개 + 안내 문구", "MVP", "Must", "US-021"),
    ("TODO-010", "할일", "공개 할 일 수락 흐름", "수락해야 본인 알림 등록. 미수락은 표시만", "MVP", "Must", "US-022"),
    ("TODO-006", "할일", "캘린더 .ics 내보내기", "외부 캘린더 앱 연동", "일부포함", "Should", "US-008"),
    ("TODO-007", "할일", "채용 마감일 자동 추출", "AI가 채용 공고 마감일 파싱 → 할 일 제안", "보류", "Could", "US-009"),
    ("TODO-008", "할일", "AI 할 일 추천", "본문 분석 → 할 일 제안", "보류", "Could", "US-007"),
    ("QR-001", "링크", "카메라 QR 스캔 (모바일)", "카메라로 QR 인식 → URL 추출", "MVP", "Must", "US-005"),
    ("QR-002", "링크", "갤러리 QR 인식", "사진첩 이미지에서 QR 추출", "보류", "Could", "US-005"),
    ("FOLDER-001", "폴더", "폴더 생성/수정/삭제", "이모지+이름. 삭제 시 링크 미분류로", "MVP", "Must", "US-003"),
    ("FOLDER-002", "폴더", "폴더 이동", "링크 상세에서 폴더 변경", "MVP", "Must", "US-003"),
    ("HOME-001", "홈", "홈 피드", "최신순 무한 스크롤, Empty State", "MVP", "Must", "US-001"),
    ("SEARCH-001", "검색", "키워드 검색", "제목·요약·태그 통합 검색, 디바운스", "MVP", "Must", "US-004"),
    ("WEB-001", "플랫폼", "데스크톱 웹 레이아웃", "1024px+ 사이드바·다중컬럼·키보드 단축키", "MVP", "Must", "US-006"),
    ("SYNC-001", "플랫폼", "모바일↔웹 자동 동기화", "한 계정 = 한 데이터, Realtime 1~2초", "MVP", "Must", "US-006"),
    ("IMPORT-001", "플랫폼", "크롬 즐겨찾기 가져오기", "bookmarks.html 업로드 → 일괄 import + AI 요약", "일부포함", "Should", "US-007"),
    ("IMPORT-002", "플랫폼", "Pocket/Raindrop import", "타 서비스 export 파일 import", "보류", "Could", "US-007"),
    ("SHARE-001", "공유 폴더", "공유 폴더 만들기", "shared 모드 + 초대 토큰", "일부포함", "Should", "US-008"),
    ("SHARE-002", "공유 폴더", "공유 폴더 초대 수락", "토큰 검증 + 멤버 가입", "일부포함", "Should", "US-008"),
    ("SHARE-003", "공유 폴더", "공유 폴더 멤버 관리", "역할·제거·나가기", "일부포함", "Should", "US-008"),
    ("VIEW-001", "콘텐츠", "YouTube 임베드 미리보기", "URL 자동 인식 → 상세 임베드", "일부포함", "Should", "US-009"),
    ("SETTINGS-001", "설정", "AI 제공자 선택", "Gemini/ChatGPT/Claude/Perplexity 등", "MVP", "Must", "US-002"),
    ("SETTINGS-002", "설정", "프로필 변경", "닉네임·아바타 이모지", "MVP", "Must", "US-001"),
    ("SETTINGS-003", "설정", "비밀번호 변경", "현재 PW 검증 후 변경", "MVP", "Must", "US-001"),
    ("SETTINGS-004", "설정", "데이터 초기화", "본인 데이터 전체 삭제", "MVP", "Must", "US-001"),
    ("THEME-001", "설정", "다크/라이트 + 강조색", "다크/라이트 + 강조색 3종(핑크/블루/그레이)", "일부포함", "Should", "US-012"),
    ("AI-001", "AI", "Chrome 확장 (외부 AI 세션)", "사용자 웹 AI 세션 활용", "보류", "Could", "US-013"),
    ("BACKUP-001", "설정", "데이터 내보내기", "JSON/CSV 다운로드", "보류", "Could", "US-014"),
    ("HIGHLIGHT-001", "콘텐츠", "본문 하이라이트", "텍스트 일부 강조 저장", "보류", "Could", ""),
    ("OFFLINE-001", "PWA", "풀 오프라인 모드", "네트워크 없이 모든 기능", "미포함", "Won't", "US-015"),
    ("NATIVE-001", "PWA", "네이티브 앱", "iOS/Android 스토어", "미포함", "Won't", "US-016"),
    ("BILLING-001", "계정", "유료 플랜", "결제·구독", "미포함", "Won't", ""),
    ("POLICY-AUTH", "정책", "토큰 만료/갱신", "JWT 액세스 15분 + 리프레시 30일", "MVP", "Must", ""),
    ("POLICY-AI", "정책", "AI 요약 호출", "외부 세션→API→mock 폴백 사다리", "MVP", "Must", ""),
    ("POLICY-LOCAL", "정책", "로컬·서버 저장 분리", "게스트 로컬 / 정식 서버+캐시, 오프라인 큐잉", "MVP", "Must", ""),
    ("POLICY-SHARED", "정책", "공유 폴더 권한", "owner/member/viewer + 토큰 7일", "일부포함", "Should", ""),
]

# ===== Sheet 2: 기능명세_상세 — 기능별 detail rows =====
# 각 기능 = (기능ID, 화면ID리스트, 섹션, 기능명, 우선순위, API연동, 포함여부, 상태, 메모, [(항목, 내용), ...])
DETAIL_FEATURES = [
    # ----- AUTH-001 -----
    ("AUTH-001", ["#AUTH-LOGIN", "#AUTH-SIGNUP", "#AUTH-EMAIL-VERIFY"], "인증", "이메일 회원가입", "Must",
     "회원가입 API\n이메일 인증 코드 발송 API\n인증 코드 검증 API\n프로필 생성 트리거", "MVP", "기획완료",
     "결정: 이메일 6자리 코드. SMS는 v1.x 검토",
     [
        ("설명", "이메일·비밀번호·닉네임으로 계정 생성. 이메일 인증코드 검증 후 자동 로그인.\n\n진입점 A: #AUTH-LOGIN의 \"회원가입\" 탭\n진입점 B: 게스트 모드 헤더 \"가입하기\" 버튼 (FLOW-003)"),
        ("입력", "- 이메일: 형식 검증 + 중복 불가\n- 비밀번호: 8~64자, 영문 + 숫자 필수, 특수문자 가능\n- 닉네임: 1~20자, 한글/영문/숫자\n- 인증코드: 6자리 숫자"),
        ("출력", "- 인증 메일 발송\n- 인증 성공 시 JWT 액세스+리프레시 토큰 발급\n- profiles 자동 생성 후 #HOME"),
        ("정책", "[기본값] 아바타: 🌸\n[시간] 인증코드 유효 10분\n[재시도] 재발송 시 이전 코드 무효\n[처리 순서] 인증 완료 전까지 가입 진행 불가\n[동기화] 가입 완료 즉시 모든 디바이스 로그인 가능 (POLICY-AUTH 참고)"),
        ("예외", "- 이메일 형식 오류: '올바른 이메일 형식이 아닙니다'\n- 중복 이메일: '이미 사용 중인 이메일입니다' + 로그인 탭 유도\n- 비밀번호 정책 미충족: '8자 이상, 영문+숫자 포함이 필요해요'\n- 인증코드 불일치: '인증코드가 일치하지 않아요'\n- 인증코드 만료: '인증 시간이 만료됐어요. 재발송해주세요'\n- profiles 자동 생성 실패: 1초 대기 후 1회 재시도 → 실패 시 로그아웃 + 안내\n- 네트워크 오류: '네트워크 연결을 확인해주세요'"),
    ]),
    # ----- AUTH-002 -----
    ("AUTH-002", ["#AUTH-LOGIN"], "인증", "이메일 로그인", "Must",
     "로그인 API", "MVP", "기획완료", "",
     [
        ("설명", "이메일·비밀번호로 로그인. 성공 시 액세스+리프레시 토큰 발급, 자동 로그인 유지"),
        ("입력", "- 이메일\n- 비밀번호\n- (자동) 디바이스 ID, FCM 토큰"),
        ("출력", "- JWT 토큰 페어\n- 사용자 프로필\n- #HOME 진입"),
        ("정책", "[횟수 제한] 5회 연속 실패 시 잠시 시도 차단 (TBD: 시간)\n[저장] 토큰은 secure storage / httpOnly 쿠키\n[동기화] 로그인 즉시 SYNC-001 활성"),
        ("예외", "- 잘못된 자격증명: '이메일 또는 비밀번호가 일치하지 않아요' (어떤 항목이 틀렸는지 표기 안 함, 보안)\n- 미가입 이메일: 동일 메시지\n- 5회 실패: '잠시 후 다시 시도해주세요'\n- 네트워크 오류: 토스트 + 게스트 모드 유도"),
    ]),
    # ----- AUTH-003 -----
    ("AUTH-003", ["#AUTH-LOGIN", "#HOME"], "인증", "게스트 모드 (프로토타입 전용)", "Won't",
     "(클라이언트, feature flag)", "프로토타입", "기획완료",
     "결정: 출시 시 제거. 마이그레이션 로직 구현 불필요(게스트 데이터 폐기)",
     [
        ("설명", "(프로토타입 전용) 가입 없이 샘플 데이터로 기능 체험. 출시 빌드에서는 \"둘러보기\" 버튼 자체 비노출 (feature flag로 제어)"),
        ("정책", "[활성 조건] feature flag GUEST_MODE=true 일 때만 노출 / [출시] GUEST_MODE=false 빌드로 배포"),
    ]),
    # ----- AUTH-004 -----
    ("AUTH-004", ["#SPLASH"], "인증", "자동 로그인 / 토큰 갱신", "Must",
     "토큰 갱신 API", "MVP", "기획완료", "",
     [
        ("설명", "앱·웹 진입 시 저장된 토큰으로 자동 로그인. 액세스 토큰 만료 시 리프레시 토큰으로 갱신, 둘 다 만료 시 로그아웃"),
        ("입력", "(자동) 저장된 액세스·리프레시 토큰"),
        ("출력", "- 토큰 유효: #HOME 즉시\n- 액세스 만료 + 리프레시 OK: 새 토큰 저장 후 #HOME\n- 리프레시 만료: #AUTH-LOGIN"),
        ("정책", "POLICY-AUTH 참고\n[처리 순서] 401 응답 시 자동 1회 갱신 후 원래 요청 재시도\n[동기화] 갱신은 모든 활성 디바이스가 독립적으로 처리"),
        ("예외", "- 갱신 API 실패 (네트워크): 캐시된 #HOME 표시 + \"곧 다시 연결할게요\" 토스트\n- 리프레시 토큰 무효(서버 invalidate): 강제 로그아웃 + #AUTH-LOGIN"),
    ]),
    # ----- AUTH-005 (보류 — 빈 행) -----
    ("AUTH-005", ["#AUTH-LOGIN"], "인증", "소셜 로그인", "Should",
     "", "보류", "기획중", "MVP 이후 v1.x 추가",
     [("설명", "Google/Apple/Kakao OAuth 가입/로그인. v1.x 추가 예정 — 보류")]),
    # ----- LINK-001 -----
    ("LINK-001", ["#HOME", "#LINK-ADD"], "링크", "URL 직접 입력 추가", "Must",
     "링크 생성 API\nAI 요약 큐 publish", "MVP", "기획완료", "",
     [
        ("설명", "사용자가 URL을 직접 입력 또는 붙여넣어 저장하는 1차 진입 경로. + 버튼 → 시트 슬라이드 업 → 입력 → 저장. 저장 직후 LINK-004(AI 요약) 자동 트리거"),
        ("입력", "- URL (필수, 2048자 이하, http/https)\n- 폴더 (선택, 기본=미분류)\n- 메모 (선택, 200자)\n- (자동) source_type='url', owner_id"),
        ("출력", "- links INSERT → 홈 피드 최상단에 카드 fade-in\n- 토스트 \"저장됐어요\"\n- LINK-004 비동기 트리거"),
        ("정책", "[기본값] 폴더=미분류\n[제약] URL 길이 ≤ 2048자, 한 유저당 ≤ 10,000개\n[중복 처리] 동일 owner의 동일 URL은 INSERT 차단\n[동기화] INSERT 직후 SYNC-001로 다른 디바이스에 1~2초 내 반영"),
        ("예외", "- URL 형식 오류: '올바른 URL 형식이 아니에요' + 저장 버튼 비활성\n- 중복 URL: 모달 \"이미 저장된 링크예요. 상세로 이동할까요?\" → \"이동\" 또는 \"취소\"\n- 한도 초과 (10,000개): '저장 한도에 도달했어요. 오래된 링크를 정리해주세요'\n- 네트워크 오류: 로컬 큐잉 + '곧 다시 시도할게요' (POLICY-LOCAL 참고)"),
    ]),
    # ----- LINK-002 -----
    ("LINK-002", ["#LINK-ADD"], "링크", "클립보드 자동 감지", "Must",
     "(클라이언트)", "MVP", "기획완료", "",
     [
        ("설명", "링크 추가 시트 진입 즉시 navigator.clipboard로 텍스트 읽고, URL이면 입력창에 자동 채움"),
        ("입력", "(자동) 클립보드 텍스트"),
        ("출력", "- URL이면 입력창 자동 채움 + \"클립보드에서 가져왔어요\" 미세 안내\n- 아니면 빈 입력창"),
        ("정책", "[권한] 권한 거부 시 자동 채움 영구 비활성\n[정규식] http(s):// 시작 또는 도메인 패턴 인식"),
        ("예외", "- 권한 미허용: 자동 채움 생략, 안내 없음\n- 클립보드 비어있음/텍스트 아님: 빈 상태\n- 매우 긴 텍스트(2000자+): URL 추출 실패, 빈 상태"),
    ]),
    # ----- LINK-003 -----
    ("LINK-003", ["#SHARE-RECEIVE", "#LINK-ADD"], "링크", "시스템 공유 시트 수신", "Must",
     "(manifest 설정)", "MVP", "기획완료", "",
     [
        ("설명", "manifest.json의 share_target으로 외부 앱(카톡/인스타/사파리/크롬)이 OS 공유 메뉴에 \"QLINK\" 노출. 사용자가 선택 시 PWA가 URL 받아 #LINK-ADD에 자동 채움"),
        ("입력", "(외부) URL 또는 텍스트 (text/plain, text/url)"),
        ("출력", "- URL 자동 채움 → 사용자가 \"저장\" 누르면 LINK-001 진행"),
        ("정책", "[처리 순서] 미로그인 시 토큰 임시 보관 → 로그인 후 자동 #LINK-ADD 진입\n[정규식] 텍스트 페이로드에서 URL 정규식 추출"),
        ("예외", "- URL 없음: 빈 #LINK-ADD 진입 + \"공유받은 텍스트에서 링크를 못 찾았어요\"\n- 게스트: 로컬 저장 + \"가입하면 영구 보관돼요\" 배너\n- 중복 URL: LINK-001 분기 동일"),
    ]),
    # ----- LINK-004 -----
    ("LINK-004", ["#LINK-ADD", "#LINK-DETAIL"], "링크", "AI 한 줄 요약 + 자동 태그", "Must",
     "AI 요약 API (provider별 어댑터)\nai_jobs INSERT/UPDATE", "MVP", "기획완료",
     "결정: MVP=mock+사용자 API키 / v1.x=외부 세션 도입. 단건은 외부 세션 미사용, 일괄 import만 외부 세션 우선 검토",
     [
        ("설명", "링크 저장 직후 비동기로 AI 요약 호출. 사다리 폴백: (1) Chrome 확장 외부 세션(Could) → (2) 사용자 본인 API 키 → (3) 서버 mock. 응답 받으면 links UPDATE + Realtime publish. POLICY-AI 참고"),
        ("입력", "- link_id\n- URL\n- (옵션) 페이지 메타(og:title, og:description)\n- 사용자 선택 provider (SETTINGS-001)"),
        ("출력", "- one_liner: 한 줄 요약 (200자 이내)\n- tags: 자동 태그 3개\n- links UPDATE + ai_jobs 레코드\n- UI에 요약 텍스트 fade-in"),
        ("정책", "[기본값] provider=SETTINGS-001 사용자 선택값, 미설정 시 'mock'\n[시간] 응답 1.4초 이내 (timeout 후 fallback)\n[재시도] 1회 자동 재시도(2초 후), 실패 시 빈 요약으로 저장\n[데이터·채점] 응답 raw는 ai_jobs.raw_response에 jsonb로 보존\n[동기화] 요약 완료 즉시 SYNC-001로 다른 디바이스 반영"),
        ("예외", "- 모든 폴백 실패: 빈 요약 + 상세에 \"다시 요약\" 버튼\n- 응답 스키마 불일치(JSON 파싱 실패): 'AI 응답을 이해하지 못했어요' + 재시도 옵션\n- URL 비공개/메타 없음: URL 자체로 추측 요약\n- 한도 초과(외부 API quota): provider 자동 폴백 + 사용자 안내"),
    ]),
    # ----- LINK-005 -----
    ("LINK-005", ["#LINK-DETAIL"], "링크", "링크 상세 (열기·공유·이동·삭제)", "Must",
     "링크 조회 API\n링크 삭제 API\n링크 폴더 이동 API", "MVP", "기획완료", "",
     [
        ("설명", "링크 카드 탭 → 시트(모바일)·사이드패널(웹)으로 상세 노출. 액션 5개: 열기 / 공유 / 폴더 이동 / 알림 설정 / 삭제"),
        ("입력", "(탭) link_id\n(액션) 사용자 선택"),
        ("출력", "- 열기: 새 탭/외부 브라우저로 URL 열기\n- 공유: OS 공유 시트\n- 폴더 이동: 폴더 픽커 → UPDATE\n- 알림: LINK-007\n- 삭제: 확인 모달 → DELETE"),
        ("정책", "[차별 인터랙션] 모바일=시트 슬라이드 업, 웹=우측 사이드패널 (WEB-001)\n[확인] 삭제는 항상 모달 한 번\n[동기화] 모든 액션은 SYNC-001로 즉시 반영"),
        ("예외", "- 삭제 후 다른 디바이스에서 같은 링크 보는 중: Realtime으로 카드 fade-out + \"다른 디바이스에서 삭제됐어요\"\n- 권한 없음(공유 폴더 viewer): 삭제 버튼 비활성, 툴팁 \"읽기 권한이에요\""),
    ]),
    # ----- LINK-006 -----
    ("LINK-006", ["#LINK-EDIT"], "링크", "링크 메모/제목 수정", "Must",
     "링크 수정 API", "MVP", "기획완료", "",
     [
        ("설명", "사용자가 제목·요약·태그를 직접 편집. AI 요약이 마음에 안 들 때 수동 수정"),
        ("입력", "- 제목 (200자)\n- 요약 (500자)\n- 태그 (5개, 각 20자)"),
        ("출력", "- links UPDATE + Realtime publish"),
        ("정책", "[제약] 길이 한도 초과 시 인풋 카운터 빨간색 + 저장 비활성\n[동시 편집] last-write-wins (POLICY-LOCAL 참고)"),
        ("예외", "- 빈 제목 저장 시도: '제목을 입력해주세요'\n- 태그 6개 이상: 입력 자체 차단\n- 다른 디바이스에서 같은 링크 동시 편집: 저장 시 서버 timestamp 비교, 충돌 시 토스트 \"다른 디바이스에서 먼저 저장됐어요. 다시 확인해주세요\""),
    ]),
    # ----- LINK-007 -----
    ("LINK-007", [], "링크", "(DEPRECATED) 단일 알림 예약", "Won't", "", "미포함", "기획중",
     "v1.1에서 TODO-001~005로 통합. 단일 reminder_at 모델 폐기, 다중 todos 모델 채택",
     [("설명", "기존 단일 reminder_at 모델은 TODO-001~005에 통합되어 폐기됨")]),
    ("TODO-001", ["#LINK-DETAIL", "#TODO-CREATE", "#TODO-EDIT"], "할일", "링크당 다중 할 일", "Must",
     "POST/PATCH/DELETE /links/:id/todos", "MVP", "기획완료", "링크당 ≤20. 추가 시 EventBridge Rule 등록",
     [
        ("설명", "한 링크에 여러 개 할 일을 매달 수 있다. 각 할 일은 제목 + 3가지 알림 모드(없음/시간선택/반복알림) 중 하나. 링크당 최대 20개. UI: 편집 시트에 할 일 행 N개 + \"＋ 할 일 추가\" 버튼"),
        ("입력", "- 제목 (1~50자, 필수)\n- 알림 모드: 없음 / 시간선택 / 반복알림 (chip 3개 중 1개)\n- 시간선택: 캘린더 picker(날짜) + 시간 picker(HH:MM, KST)\n- 반복알림: 요일 체크박스(일·월·화·수·목·금·토, default 모두 선택=매일) + 시간 + 종료일(선택)"),
        ("출력", "link_todos INSERT/UPDATE/DELETE + Realtime publish (TODO_CREATED/UPDATED/DELETED)"),
        ("정책", "[기본값] 알림 모드=없음, 반복 요일=매일(7개)\n[제약] 링크당 20개, 제목 50자, 반복 종료일 ≤ 시작+365일\n[정렬] sort_order 기준\n[동기화] SYNC-001 즉시 반영"),
        ("예외", "- 빈 제목: '할 일 제목을 입력해주세요'\n- 50자 초과: 입력 차단\n- 한도 20개 초과: '한 링크에 할 일은 20개까지 만들 수 있어요'\n- 시간선택 모드 + 과거 시각: picker 차단\n- 반복 모드 + 요일 0개: '하나 이상의 요일을 선택해주세요'\n- 권한 미허용: '인앱 토스트로만 알려드려요'"),
    ]),
    ("TODO-002", ["#TODO-CREATE", "#TODO-EDIT"], "할일", "반복 알림 (요일 기반)", "Must",
     "POST /links/:id/todos (notifyMode='recurring')", "MVP", "기획완료", "default 매일=7개 모두 선택",
     [
        ("설명", "매일/평일/주말/사용자 정의 요일 + 시간으로 반복 알림. 요일 체크박스 7개 (일·월·화·수·목·금·토). default 모두 체크(매일). 사용자가 해제해서 패턴 만들기"),
        ("입력", "weekdays: int[](0=일~6=토), notifyTime: HH:MM(KST), endDate: YYYY-MM-DD(선택)"),
        ("출력", "EventBridge Rule 등록 (다음 회차 트리거)"),
        ("정책", "[기본값] weekdays=[0,1,2,3,4,5,6], endDate=null(시스템 365일 cap)\n[표시] '매일 21:00' / '평일 21:00' / '주말 10:00' / '월수금 21:00' 자연 한국어\n[다음 회차] 현재 시각 이후 weekdays 중 가장 가까운 시각\n[발송] FCM Topic + 사용자 채널 토글 적용"),
        ("예외", "- 요일 0개: '하나 이상의 요일을 선택해주세요'\n- 시간 미입력: '시간을 입력해주세요'\n- endDate 도달: Rule 자동 삭제, 할 일 status='완료'"),
    ]),
    ("TODO-003", ["#TODO-LIST", "#LINK-DETAIL"], "할일", "할 일 완료 체크 (회차별 + 누적)", "Must",
     "POST /todos/:id/complete\nGET /todos/:id/occurrences", "MVP", "기획완료",
     "결정(TBD-13): 회차별 완료 채택. 매주 강의 → 이번 주 ✓ → 다음 주 또 활성. 누적 이력은 토글 접기/펼치기",
     [
        ("설명", "반복 모드는 회차별 완료(이번 주 완료해도 다음 주에 또 활성). 1회/시간없음 모드는 todo 자체 완료. 누적 완료 이력은 토글로 접기·펼치기 (default 접힘). \"✓ 완료한 N회 · 이번 달 M회 ▼\" 형태로 요약, 펼치면 회차별 완료 시각 노출 (max-height 200px + 스크롤)"),
        ("입력", "todoId, completed: bool, (반복 모드) occurrenceDate: 'YYYY-MM-DD' (생략 시 오늘 KST)"),
        ("출력", "- 반복: todo_occurrences INSERT/DELETE\n- 그 외: link_todos.completed_at UPDATE\n- Realtime TODO_UPDATED publish"),
        ("정책", "[회차별 완료] 반복 todo는 오늘 회차만 토글. 토글 후 카드의 \"활성 회차\"가 다음 예정일로 자동 이동\n[누적 표시] 토글 default 접힘. 펼치면 시간 역순\n[낙관적 UI] 즉시 ✓ 후 UPDATE, 실패 시 롤백\n[반복 종료] 사용자가 명시 액션 시에만 EventBridge Rule 일괄 제거"),
        ("예외", "- 네트워크 오류: 로컬 큐잉, 재연결 시 동기화\n- 동시 토글: last-write-wins\n- 같은 날짜 중복: PK (todo_id, occurrence_date) 자동 dedupe"),
    ]),
    ("TODO-004", ["#TODO-LIST"], "할일", "할일 통합 화면", "Must",
     "GET /todos?filter=...", "MVP", "기획완료", "탭바 5번째",
     [
        ("설명", "모바일 탭바 \"할일\" / 데스크톱 사이드바. 모든 link_todos 평탄화 노출. 카드=체크박스+제목+알림 시각/반복 배지+원본 링크 미리보기"),
        ("입력", "필터 chip(전체/미완료/알림예정/기간지남/완료) + 정렬"),
        ("출력", "할 일 카드 리스트"),
        ("정책", "[정렬] 미완료 우선 → 알림 임박순 → 최신 추가순. 완료는 맨 아래\n[페이지네이션] cursor 기반 30개\n[차별 인터랙션] 데스크톱 단축키 J/K/X/Enter"),
        ("예외", "- 빈 결과: '이 필터에 해당하는 할 일이 없어요' Empty State"),
    ]),
    ("TODO-005", [], "할일", "할 일 알림 발송 (FCM)", "Must",
     "EventBridge Rule + Lambda + FCM", "MVP", "기획완료", "DLQ 모니터링, 24h 재시도",
     [
        ("설명", "알림 시각에 EventBridge Rule이 Lambda 트리거 → FCM Topic 발송. 사용자 활성 디바이스 모두에 푸시 (모바일/PC 채널 토글 적용)"),
        ("입력", "(시스템) Rule 발화"),
        ("출력", "FCM 발송 + WebSocket TODO_FIRED 이벤트"),
        ("정책", "[신뢰성] DLQ 모니터링, 24시간 후 1회 재시도\n[반복 알림] 발송 직후 다음 회차 Rule 등록 (재귀 스케줄)\n[종료] endDate 또는 완료 시 Rule 자동 삭제"),
        ("예외", "- 권한 미허용: 인앱 토스트 폴백\n- 두 채널 모두 OFF: 발송 스킵\n- 링크 삭제됨: 발송 안 함"),
    ]),
    ("TODO-009", ["#TODO-CREATE", "#TODO-EDIT", "#LINK-DETAIL"], "할일", "할 일 가시성 (공유 폴더)", "Must",
     "POST /links/:id/todos (visibility)\nPATCH /todos/:id (visibility)", "MVP", "기획완료",
     "메모는 항상 private + 안내 문구. 공유 폴더에서만 가시성 chip 노출",
     [
        ("설명", "공유 폴더 link의 할 일은 🔒 나만 보기(default) / 👥 공유자에게 공유 선택. 개인 폴더는 옵션 비노출, 자동 private. 메모는 가시성 옵션 없음, 항상 작성자 전용 + '🔒 나만 볼 수 있는 내용입니다' 안내"),
        ("입력", "visibility: 'private' | 'public'"),
        ("출력", "link_todos.visibility UPDATE + 공개로 변경 시 멤버에게 Realtime 노출"),
        ("정책", "[기본값] private\n[개인 폴더] visibility 자동 private, UI에 노출 안 함\n[메모] 가시성 변경 불가\n[작성자 표시] public todo는 멤버 화면에 '👥 {닉네임}이 공유한 할 일' 라벨"),
        ("예외", "- 비멤버 공유 todo 조회: 403\n- 메모를 공개로 변경 시도: API에서 차단"),
    ]),
    ("TODO-010", ["#TODO-LIST", "#LINK-DETAIL"], "할일", "공개 할 일 수락 흐름", "Must",
     "POST /todos/:id/accept\nDELETE /todos/:id/accept\nGET /todos/:id/acceptances", "MVP", "기획완료",
     "의도치 않은 알림 폭탄 방지. 작성자에게 누가 수락했는지 카운터 표시",
     [
        ("설명", "공유 폴더에서 멤버 A가 public todo 등록 → 멤버 B는 화면에서 보지만 '내 할 일에 추가' 수락해야 본인 알림으로 등록. 미수락은 표시만 (정보 공유 수준)"),
        ("입력", "todoId (수락 또는 취소)"),
        ("출력", "todo_acceptances INSERT/DELETE + WebSocket TODO_ACCEPTED"),
        ("정책", "[수락 UI] 공개 todo 카드에 '내 할 일에 추가' 버튼. 수락 후 '✓ 추가됨'\n[작성자] '👥 N명 수락' 카운터\n[알림 발송] 작성자 + 수락 멤버에게만 FCM\n[회차별 완료] 수락 멤버는 본인 occurrences 따로 관리"),
        ("예외", "- 본인이 작성자: 수락 버튼 비활성\n- 비멤버: 403\n- 이미 수락: '이미 추가됨'"),
    ]),
    ("TODO-006", [], "할일", "캘린더 .ics 내보내기", "Should", "", "일부포함", "기획중",
     "외부 캘린더(Google/Apple/네이버) 연동",
     [("설명", "할 일 → .ics 파일로 다운로드. 외부 캘린더 import. 반복 알림은 RRULE로 변환")]),
    ("TODO-007", [], "할일", "채용 마감일 자동 추출", "Could", "", "보류", "기획중",
     "v1.x — 채용 도메인(잡코리아/사람인/원티드/링커리어) 패턴 학습",
     [("설명", "채용 공고 도메인 인식 → AI가 마감일 파싱 → 할 일 제안 (D-3 자소서, D-1 포트폴리오, 당일 지원)")]),
    ("TODO-008", [], "할일", "AI 할 일 추천", "Could", "", "보류", "기획중", "본문 요약 → 할 일 후보 추출",
     [("설명", "링크 본문 분석 → 액션 아이템 추출 → 할 일 후보로 사용자에게 제안")]),
    # ----- QR-001 -----
    ("QR-001", ["#LINK-ADD", "#QR-SCAN"], "링크", "카메라 QR 스캔 (모바일)", "Must",
     "(클라이언트, html5-qrcode)", "MVP", "기획완료", "",
     [
        ("설명", "html5-qrcode로 카메라 스트림 인식. URL 형태의 QR이면 #LINK-ADD에 자동 채움. 데스크톱 웹에서는 비활성"),
        ("입력", "(자동) MediaStream\n사용자: QR 비추기"),
        ("출력", "- URL 추출 → #LINK-ADD 입력창 채움"),
        ("정책", "[권한] 카메라 권한 거부 시 URL 직접 입력 폴백 + 안내\n[환경] HTTPS 또는 localhost 필수\n[차별 인터랙션] 모바일 전용. 웹에서는 QR 탭 비활성 + 툴팁 \"모바일에서 사용 가능해요\""),
        ("예외", "- 권한 거부: '카메라 권한이 필요해요. 설정에서 허용해주세요'\n- HTTPS 아님: 'QR 스캔은 보안 연결에서만 동작해요'\n- 인식 실패 10초: '다시 비춰주세요' + Mock 입력(개발용)\n- URL이 아닌 QR: 'URL이 아닌 QR이에요'"),
    ]),
    # ----- QR-002 (보류 빈 행) -----
    ("QR-002", [], "링크", "갤러리 QR 인식", "Could", "", "보류", "기획중", "v1.x 추가 예정",
     [("설명", "사진첩 이미지에서 QR 추출 (Media Capture + jsQR)")]),
    # ----- FOLDER-001 -----
    ("FOLDER-001", ["#FOLDER-LIST", "#FOLDER-CREATE", "#FOLDER-DETAIL"], "폴더", "폴더 생성/수정/삭제", "Must",
     "폴더 CRUD API", "MVP", "기획완료", "",
     [
        ("설명", "이모지 + 이름으로 폴더 관리. 시스템 폴더 \"미분류\"는 삭제 불가. 공유 폴더는 SHARE-001 참고"),
        ("입력", "(생성) 이름 1~30자, 이모지\n(수정) 이름·이모지 변경\n(삭제) 폴더 ID + 확인"),
        ("출력", "folders INSERT/UPDATE/DELETE + Realtime"),
        ("정책", "[기본값] 이모지=📁\n[제약] 폴더 100개 한도\n[삭제 처리] 폴더 삭제 시 소속 링크는 folder_id=NULL (미분류로 자동 이동)\n[공유 폴더] owner만 이름·이모지 변경 가능"),
        ("예외", "- 빈 이름: '폴더 이름을 입력해주세요'\n- 30자 초과: 입력 자체 차단\n- 공유 폴더 멤버가 삭제 시도: '오너만 삭제할 수 있어요' + \"나가기\" 버튼\n- 미분류 삭제 시도: 메뉴에 삭제 노출 안 함\n- 100개 한도: '폴더는 최대 100개까지'"),
    ]),
    # ----- FOLDER-002 -----
    ("FOLDER-002", ["#LINK-DETAIL", "#FOLDER-PICKER"], "폴더", "폴더 이동", "Must",
     "링크 폴더 이동 API", "MVP", "기획완료", "",
     [
        ("설명", "링크 상세에서 폴더 변경. 픽커 모달에 사용자 폴더 + 미분류 노출"),
        ("입력", "link_id, target_folder_id"),
        ("출력", "links.folder_id UPDATE + Realtime"),
        ("정책", "[동기화] 다른 디바이스에서 카드가 다른 폴더로 부드럽게 이동"),
        ("예외", "- 공유 폴더(viewer 권한)로 이동 시도: '읽기 권한 폴더에는 옮길 수 없어요'\n- 삭제된 폴더 ID: 미분류로 자동 fallback"),
    ]),
    # ----- HOME-001 -----
    ("HOME-001", ["#HOME", "#HOME-EMPTY"], "홈", "홈 피드", "Must",
     "링크 목록 API (cursor)", "MVP", "기획완료",
     "TBD: 사용자 설정으로 인기/오래된순 추가?",
     [
        ("설명", "사용자의 모든 링크를 created_at desc로 무한 스크롤. 페이지당 30개. 게스트 시드 데이터 또는 Empty State"),
        ("입력", "(자동) cursor (last_created_at)"),
        ("출력", "링크 카드 리스트 / Empty State 일러스트+CTA"),
        ("정책", "[기본값] 정렬: 최신순\n[페이지네이션] cursor 기반 30개\n[차별 인터랙션] 모바일=1열, 웹 1024px+=2열, 1440px+=3열, 1920px+=4열 (WEB-001)"),
        ("예외", "- 첫 진입 + 데이터 0개: Empty State \"+ 버튼으로 첫 링크를 저장해보세요\"\n- 네트워크 오류: 캐시된 마지막 페이지 + \"오프라인이에요\" 배너"),
    ]),
    # ----- SEARCH-001 -----
    ("SEARCH-001", ["#SEARCH", "#SEARCH-EMPTY"], "검색", "키워드 검색", "Must",
     "검색 API (ILIKE → tsvector)", "MVP", "기획완료", "",
     [
        ("설명", "제목·요약·태그 통합 검색. 디바운스 200ms. 결과 카드에 키워드 하이라이트. 데스크톱 웹은 ⌘K/Ctrl+K 단축키"),
        ("입력", "키워드 (2자 이상)"),
        ("출력", "매칭 링크 카드 리스트 + 키워드 강조"),
        ("정책", "[기본값] 검색 대상=제목/요약/태그 동시\n[구현] v1: ILIKE '%kw%' (Postgres) / v2: tsvector + gin 인덱스\n[동기화] 새로 저장된 링크도 즉시 검색 가능"),
        ("예외", "- 1자: 트리거 안 함, '2자 이상 입력해주세요'\n- 결과 0개: \"'{키워드}' 결과가 없어요\" + 최근 검색어 추천\n- 네트워크 오류: 로컬 캐시 결과 + '오프라인 결과를 보여드려요'"),
    ]),
    # ----- WEB-001 -----
    ("WEB-001", ["(전 화면 반응형)"], "플랫폼", "데스크톱 웹 레이아웃 (반응형 PWA)", "Must",
     "(클라이언트 레이아웃)", "MVP", "기획완료",
     "TBD: 사이드바 폭(고정 240px vs 가변), 다크모드 색상 토큰은 THEME-001과 연동",
     [
        ("설명", "단일 React PWA 코드베이스가 화면 폭에 따라 모바일·태블릿·데스크톱 레이아웃을 자동 전환. 데스크톱은 좌측 사이드바(폴더 트리) + 메인 그리드(2~4컬럼) + 우측 상세 패널 구조"),
        ("입력", "window.innerWidth (브레이크포인트 매칭)"),
        ("출력", "- <1024px: 모바일 레이아웃(단일 컬럼 + 하단 탭바)\n- 1024~1439px: 사이드바 + 2컬럼 그리드\n- 1440~1919px: 사이드바 + 3컬럼\n- 1920px+: 사이드바 + 4컬럼 + 우측 상세 패널"),
        ("정책", "[차별 인터랙션] 데스크톱: 키보드 단축키 ⌘K(검색) / N(새 링크) / J·K(카드 이동) / Enter(상세) / Del(삭제)\n[차별 인터랙션] 데스크톱: 카드 hover 시 액션 버튼 노출\n[모바일 전용] QR-001은 데스크톱에서 비활성"),
        ("예외", "- 화면 폭 변경(창 리사이즈): 즉시 레이아웃 재계산, 스크롤 위치 보존\n- PWA 설치 데스크톱: 단축키 충돌 회피(브라우저 단축키 존중)"),
    ]),
    # ----- SYNC-001 -----
    ("SYNC-001", ["(전 화면 백그라운드)"], "플랫폼", "모바일↔웹 자동 동기화", "Must",
     "AWS API Gateway WebSocket\nsync delta API", "MVP", "기획완료",
     "결정: AWS API Gateway WebSocket 채택 (단순·비용 효율). 끊김 시 5s polling 폴백. last-write-wins. CRDT는 v2 백로그",
     [
        ("설명", "한 계정 = 단일 백엔드(AWS RDS) → 디바이스 간 데이터 자동 동기화. AWS API Gateway WebSocket으로 1~2초 내 반영. 멀티탭/멀티디바이스 동시 사용 가정"),
        ("입력", "(자동) 사용자 인증 + 활성 세션"),
        ("출력", "- 각 디바이스가 links/folders/folder_members 변경 이벤트 구독\n- 변경 발생 → 모든 활성 디바이스에 publish → UI에 fade 애니메이션"),
        ("정책", "[기본] AWS API Gateway WebSocket (재연결 시 backoff 1s,2s,4s,8s,30s)\n[폴백] 1차 5s polling, 2차 60s polling\n[충돌] last-write-wins by `updated_at` (CRDT는 v2 백로그)\n[낙관적 UI] 클라이언트는 즉시 적용, 서버 응답으로 검증\n[동시 편집] 같은 필드 동시 편집 시 토스트 알림"),
        ("예외", "- WebSocket 끊김: 자동 재연결 + 마지막 sync timestamp 이후 변경분 일괄 fetch\n- 디바이스 시계 어긋남: 서버 timestamp만 사용\n- 게스트 모드(프로토타입): SYNC 비활성\n- 데이터 초기화(SETTINGS-004): 모든 디바이스에서 즉시 빈 상태"),
    ]),
    # ----- IMPORT-001 -----
    ("IMPORT-001", ["#SETTINGS-IMPORT", "#IMPORT-UPLOAD", "#IMPORT-PREVIEW", "#IMPORT-PROGRESS"], "플랫폼", "크롬 즐겨찾기 가져오기 (웹 전용)", "Should",
     "북마크 import API (배치)\nAI 요약 큐 publish", "일부포함", "기획완료",
     "결정: MVP는 비용 동의 모달 미표시 (mock·API키만 가정, 사용자가 본인 API 키 사용 시 본인 부담). 외부 세션 도입(v1.x) 시 모달 추가",
     [
        ("설명", "크롬 북마크 관리자 → 내보내기 → bookmarks.html 파일 업로드. 파싱 후 폴더 구조와 함께 미리보기 → 사용자가 폴더 선택/제외 → 일괄 INSERT + AI 요약 SQS 큐잉"),
        ("플로우", "(1) 안내(크롬 메뉴 → 북마크 관리자 → ⋮ → 북마크 내보내기) → (2) bookmarks.html 업로드 → (3) 파싱 + 미리보기(폴더 트리 + 링크 카운트) → (4) 폴더 선택/제외 → (5) \"가져오기\" → (6) 진행률 + 백그라운드 AI 요약"),
        ("입력", "- bookmarks.html (≤5MB, Netscape Bookmark Format)"),
        ("출력", "- 폴더 N개 INSERT\n- 링크 M개 INSERT\n- SQS에 AI 요약 작업 M개 publish\n- \"N개 가져왔어요. AI 요약은 백그라운드에서 진행 중이에요\""),
        ("정책", "[차별 인터랙션] 데스크톱 웹 전용 (모바일은 안내만)\n[중복 처리] 기본=스킵, 옵션=덮어쓰기\n[요약 전략] 일괄 import 시 요약은 백그라운드 (홈 피드는 빈 요약으로 즉시 노출, 요약 도착 시 fade-in)\n[한도] 폴더 100개·링크 10000개 한도 사전 검증, 초과 시 일부 선택 강제"),
        ("예외", "- 형식 오류: '크롬 즐겨찾기 HTML 파일만 가능해요'\n- 5MB 초과: '파일이 너무 커요. 5MB 이하로 잘라서 다시 시도해주세요'\n- 파싱 실패: '북마크 구조를 읽지 못했어요' + 다시 내보내기 안내\n- 한도 초과: 부분 선택 강제 + 안내\n- 모바일 진입: '데스크톱 웹에서 사용 가능한 기능이에요'"),
    ]),
    # ----- IMPORT-002 -----
    ("IMPORT-002", [], "플랫폼", "Pocket/Raindrop import", "Could", "", "보류", "기획중", "v1.x 추가 예정",
     [("설명", "Pocket·Raindrop·Goodlinks 등 export 파일 import")]),
    # ----- SHARE-001 -----
    ("SHARE-001", ["#FOLDER-DETAIL", "#SHARE-INVITE"], "공유 폴더", "공유 폴더 만들기", "Should",
     "초대 토큰 생성 API", "일부포함", "기획완료", "",
     [
        ("설명", "폴더의 shared 플래그를 true로 전환 → folder_invites에 토큰 INSERT → OS 공유 시트로 링크 발송"),
        ("입력", "folder_id (오너만)"),
        ("출력", "folders.shared=true / folder_invites INSERT / 공유 URL 생성 (https://qlink.app/invite/{token})"),
        ("정책", "POLICY-SHARED 참고\n[기본값] 토큰 유효 7일\n[권한] 오너만 발급\n[재발급] 새 토큰 발급 시 이전 토큰은 만료 전까지 유효"),
        ("예외", "- 비오너 시도: '오너만 공유할 수 있어요'\n- 미분류 폴더 공유 시도: '미분류 폴더는 공유할 수 없어요'"),
    ]),
    # ----- SHARE-002 -----
    ("SHARE-002", ["#SHARE-ACCEPT"], "공유 폴더", "공유 폴더 초대 수락", "Should",
     "accept_folder_invite RPC", "일부포함", "기획완료", "",
     [
        ("설명", "초대 링크 클릭 → 토큰 검증 → 사용자가 \"참여\" 클릭 → folder_members INSERT"),
        ("입력", "invite_token (URL 파라미터)"),
        ("출력", "folder_members INSERT (role='member') / 폴더 진입"),
        ("정책", "POLICY-SHARED 참고\n[처리 순서] 미로그인 시 토큰 임시 보관 → 로그인 후 자동 수락 화면\n[멤버 한도] 50명"),
        ("예외", "- 토큰 만료: '유효하지 않거나 만료된 초대 링크입니다'\n- 이미 멤버: 토스트 + 폴더 이동\n- 본인이 오너: '본인의 폴더예요'\n- 멤버 한도 50명 초과: '이 폴더는 50명까지만 참여할 수 있어요. 오너에게 문의하세요'"),
    ]),
    # ----- SHARE-003 -----
    ("SHARE-003", ["#FOLDER-MEMBERS"], "공유 폴더", "공유 폴더 멤버 관리", "Should",
     "멤버 조회/수정/삭제 API", "일부포함", "기획중", "v1.0 가벼운 형태로",
     [("설명", "멤버 목록 + 역할 변경(owner만) + 제거(owner) / 나가기(self)")]),
    # ----- VIEW-001 -----
    ("VIEW-001", ["#LINK-DETAIL"], "콘텐츠", "YouTube 임베드 미리보기", "Should",
     "(클라이언트, iframe)", "일부포함", "기획중", "",
     [("설명", "URL이 youtube.com/youtu.be 도메인이면 상세 화면에서 iframe 임베드. 자동재생 X, 사용자 클릭 시 재생")]),
    # ----- SETTINGS-001 -----
    ("SETTINGS-001", ["#SETTINGS-AI"], "설정", "AI 제공자 선택", "Must",
     "프로필 업데이트 API", "MVP", "기획완료", "",
     [
        ("설명", "6종 중 선택: ChatGPT / Gemini / Claude / Perplexity / 큐링크 기본(mock) / 사용자 API 키. SETTINGS-001 변경은 LINK-004의 provider 입력으로 즉시 반영"),
        ("입력", "provider 식별자"),
        ("출력", "profiles.ai_provider UPDATE + 토스트"),
        ("정책", "[기본값] 'mock' (앱 진입 시)\n[동기화] 모든 디바이스에 SYNC-001로 반영\n[API 키] 사용자 API 키 선택 시 키 입력 화면 추가 노출, 서버 KMS 암호화 저장"),
        ("예외", "- 외부 세션 provider 선택했지만 Chrome 확장 미설치: '확장 프로그램 설치가 필요해요' + 안내 링크\n- API 키 무효: 다음 호출 시 '키를 확인해주세요' 안내"),
    ]),
    # ----- SETTINGS-002 -----
    ("SETTINGS-002", ["#SETTINGS-PROFILE"], "설정", "프로필 변경", "Must",
     "프로필 업데이트 API", "MVP", "기획완료", "",
     [
        ("설명", "닉네임·아바타 이모지 변경"),
        ("입력", "display_name (1~20자), avatar (이모지 단일 문자)"),
        ("출력", "profiles UPDATE + 헤더 즉시 반영"),
        ("정책", "[동기화] 즉시 모든 디바이스 반영"),
        ("예외", "- 빈 닉네임: '닉네임을 입력해주세요'\n- 20자 초과: 입력 차단"),
    ]),
    # ----- SETTINGS-003 -----
    ("SETTINGS-003", ["#SETTINGS-PASSWORD"], "설정", "비밀번호 변경", "Must",
     "비밀번호 변경 API\n리프레시 토큰 무효화 API", "MVP", "기획완료",
     "결정: 변경 시 다른 디바이스 강제 로그아웃 + 화면 안내 문구 추가",
     [
        ("설명", "현재 비밀번호 검증 → 새 비밀번호 변경. 변경 후 다른 디바이스는 강제 로그아웃 (보안 우선)"),
        ("입력", "current_password, new_password (8~64자, 영문+숫자)"),
        ("출력", "비밀번호 변경 + 모든 다른 디바이스 리프레시 토큰 invalidate + 토스트 \"변경됐어요. 다른 기기에서는 다시 로그인해주세요\""),
        ("정책", "[보안] 현재 비밀번호 검증 필수\n[강제 로그아웃] 변경 즉시 서버에서 모든 다른 활성 리프레시 토큰 invalidate\n[화면 안내] 비밀번호 변경 화면 상단에 안내 카피: \"비밀번호를 바꾸면 로그인된 다른 기기들에서 자동으로 로그아웃돼요. 보안을 위한 조치예요.\""),
        ("예외", "- 현재 PW 불일치: '현재 비밀번호가 일치하지 않아요'\n- 새 PW 정책 미충족: 안내\n- 5회 실패: 잠시 차단"),
    ]),
    # ----- SETTINGS-004 -----
    ("SETTINGS-004", ["#SETTINGS-RESET"], "설정", "데이터 초기화", "Must",
     "데이터 일괄 삭제 API (트랜잭션)", "MVP", "기획완료",
     "결정: 즉시 hard delete (개인정보보호법 OK, 사용자 의도 그대로 존중)",
     [
        ("설명", "본인 데이터(폴더·링크·공유 멤버십) 전체 영구 삭제. 이중 확인 모달. 모든 디바이스에서 즉시 빈 상태"),
        ("입력", "\"초기화\" 버튼 → 모달 1 → \"정말 삭제\" → 모달 2 → 사용자 비밀번호 재입력"),
        ("출력", "DELETE folders/links/folder_members(self) + 모든 디바이스 SYNC"),
        ("정책", "[위험 액션] 이중 확인 + 비밀번호 재입력\n[복구 불가] 즉시 hard delete — DB에서 본인 데이터 영구 삭제, 복구 불가\n[고지] 모달 2에 명시: \"삭제하면 복구할 수 없어요. 그래도 진행할까요?\"\n[동기화] 즉시 모든 디바이스에서 빈 상태"),
        ("예외", "- 비밀번호 불일치: 모달 안에서 '비밀번호가 일치하지 않아요'\n- 처리 중 네트워크 오류: 트랜잭션으로 서버 측 부분 삭제 방지 (전부 실패하거나 전부 성공)"),
    ]),
    # ----- THEME-001 -----
    ("THEME-001", ["#SETTINGS-THEME"], "설정", "다크/라이트 + 강조색", "Should",
     "프로필 업데이트 API", "일부포함", "기획완료",
     "결정: 다크/라이트 모드 + 강조색 3종(핑크/블루/그레이)",
     [
        ("설명", "다크/라이트 모드 + 강조색 3종(핑크·블루·그레이). 사용자별 저장. 로그인 화면은 항상 라이트+핑크(브랜드)"),
        ("입력", "mode (`light`/`dark`), accent (`pink`/`blue`/`gray`)"),
        ("출력", "profiles.theme UPDATE + CSS 변수 즉시 반영 + SYNC"),
        ("정책", "[동기화] SYNC-001로 모든 디바이스 즉시 반영\n[기본값] 신규 가입자 light + pink"),
    ]),
    # ----- AI-001 -----
    ("AI-001", [], "AI", "Chrome 확장 (외부 AI 세션)", "Could", "", "보류", "기획중",
     "ToS 회색 영역 — 신중 검토",
     [("설명", "Chrome MV3 확장이 사용자 본인의 웹 Gemini/ChatGPT/Claude 세션에 프롬프트 주입 → 응답 파싱 → QLink 앱으로 postMessage")]),
    # ----- BACKUP-001 / HIGHLIGHT-001 -----
    ("BACKUP-001", [], "설정", "데이터 내보내기", "Could", "", "보류", "기획중", "", [("설명", "JSON·CSV 다운로드. IMPORT-001과 대칭")]),
    ("HIGHLIGHT-001", [], "콘텐츠", "본문 하이라이트", "Could", "", "보류", "기획중", "", [("설명", "링크 본문 텍스트 일부 강조 저장")]),
    # ----- Won't (자리만) -----
    ("OFFLINE-001", [], "PWA", "풀 오프라인 모드", "Won't", "", "미포함", "기획중", "이번 릴리즈 제외", [("설명", "네트워크 없이 모든 기능 동작 — 동기화 충돌 복잡도, v2 백로그")]),
    ("NATIVE-001", [], "PWA", "네이티브 앱", "Won't", "", "미포함", "기획중", "이번 릴리즈 제외", [("설명", "iOS/Android 스토어 등록 — PWA로 충분")]),
    ("BILLING-001", [], "계정", "유료 플랜", "Won't", "", "미포함", "기획중", "이번 릴리즈 제외", [("설명", "결제·구독 — MVP는 무료 우선")]),
    # ----- POLICY-* -----
    ("POLICY-AUTH", [], "정책", "토큰 만료/갱신 정책", "Must", "", "MVP", "기획완료", "공통 참조 정책",
     [
        ("설명", "모든 인증 흐름의 공통 토큰 정책. 다른 기능에서는 \"POLICY-AUTH 참고\"로 1줄 참조"),
        ("정책", "[시간] 액세스 토큰 15분, 리프레시 토큰 30일\n[처리 순서] 401 응답 시 자동 갱신 1회 후 원래 요청 재시도\n[저장] 토큰은 secure storage 또는 httpOnly 쿠키\n[무효화] 비밀번호 변경 시 모든 리프레시 토큰 무효화 옵션 (TBD)"),
        ("예외", "갱신 실패 시 강제 로그아웃 + #AUTH-LOGIN"),
    ]),
    ("POLICY-AI", [], "정책", "AI 요약 호출 정책", "Must", "", "MVP", "기획완료", "공통 참조 정책",
     [
        ("설명", "LINK-004와 IMPORT-001이 공통 사용. 사다리 폴백 정책"),
        ("정책", "[폴백 사다리] (1) Chrome 확장 외부 세션(설치된 경우) → (2) 사용자 API 키(설정된 경우) → (3) 서버 mock\n[시간] 단건 요약 1.4초 timeout, 일괄 요약 비동기 큐\n[재시도] 1회 자동, 실패 시 빈 요약으로 저장 + 수동 재시도 버튼\n[데이터·채점] raw_response는 ai_jobs 테이블에 jsonb 보존"),
        ("예외", "모든 폴백 실패 시 빈 요약. 사용자에게 토스트 안내"),
    ]),
    ("POLICY-LOCAL", [], "정책", "로컬·서버 저장 분리", "Must", "", "MVP", "기획완료", "공통 참조 정책",
     [
        ("설명", "게스트와 정식 사용자의 데이터 위치 분리. 오프라인 시 큐잉"),
        ("정책", "[로컬 저장 대상] 게스트 모드 전체 데이터 / 클립보드 자동 채움 임시값 / 화면 캐시 (PWA Service Worker)\n[서버 저장 대상] 정식 사용자의 모든 데이터\n[오프라인 큐잉] 정식 사용자가 오프라인에서 저장 시 IndexedDB 큐 → 재연결 시 자동 동기화\n[충돌] last-write-wins by updated_at"),
        ("예외", "로컬 큐 5MB 초과 시 사용자에게 안내"),
    ]),
    ("POLICY-SHARED", [], "정책", "공유 폴더 권한 정책", "Should", "", "일부포함", "기획완료", "공통 참조 정책",
     [
        ("설명", "SHARE-001/002/003 공통 권한 모델"),
        ("정책", "[역할] owner: 폴더·링크·멤버 전체 권한 / member: 링크 추가·읽기·자기 추가한 링크 삭제 / viewer: 읽기만\n[기본] 초대 수락 시 role='member'\n[토큰] 유효 7일, 1회용 아님(여러 명 사용 가능)\n[멤버 한도] 50명"),
        ("예외", "권한 부족 액션은 UI에서 비활성 + 툴팁. 백엔드는 항상 재검증 (Lambda 핸들러)"),
    ]),
]


def build_workbook(out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    # ============= Sheet 1: 기능명세_목록 =============
    ws1 = wb.create_sheet(f"기능명세_목록 v.{VERSION}")
    ws1.cell(row=1, column=1, value=f"{PROJECT_NAME} — 기능 명세 목록  |  v{VERSION}").font = TITLE_FONT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    headers = ["ID", "섹션", "기능명", "한줄 설명", "포함\n여부", "우선\n순위", "관련 US"]
    for c, h in enumerate(headers, 1):
        style_header(ws1.cell(row=2, column=c, value=h))

    widths = {1: 18, 2: 12, 3: 30, 4: 60, 5: 10, 6: 10, 7: 14}
    for col, w in widths.items():
        ws1.column_dimensions[get_column_letter(col)].width = w

    for r_idx, row in enumerate(LIST_ROWS, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP_TOP
            # 미포함/보류는 회색
            if row[4] in ("미포함", "보류"):
                cell.font = GRAY_FONT

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 32
    ws1.freeze_panes = "A3"

    # ============= Sheet 2: 기능명세_상세 =============
    ws2 = wb.create_sheet(f"기능명세_상세 v.{VERSION}")
    ws2.cell(row=1, column=1, value=f"{PROJECT_NAME} — 기능 명세 상세  | v.{VERSION}").font = TITLE_FONT
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    guide = ("[항목]\n- 설명/입력/출력/정책/예외/플로우\n"
             "- #: 케이스별 분기  - (Step1)/(A-Step2): 다단계\n"
             "- 입력: 앱 → 서버 / 출력: 서버 → 앱")
    cell = ws2.cell(row=1, column=6, value=guide)
    cell.font = GUIDE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.merge_cells(start_row=1, start_column=6, end_row=1, end_column=10)

    headers2 = ["기능ID\n(#화면ID)", "섹션", "기능명", "항목", "내용",
                "우선\n순위", "API 연동", "포함\n여부", "상태", "메모/이슈"]
    for c, h in enumerate(headers2, 1):
        style_header(ws2.cell(row=2, column=c, value=h))

    widths2 = {1: 26, 2: 12, 3: 28, 4: 18, 5: 80, 6: 8, 7: 26, 8: 10, 9: 10, 10: 36}
    for col, w in widths2.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.row_dimensions[1].height = 70
    ws2.row_dimensions[2].height = 32
    ws2.freeze_panes = "A3"

    # detail rows
    r = 3
    for feat in DETAIL_FEATURES:
        fid, screens, section, name, prio, api, include, status, memo, items = feat
        # build first-cell value: fid + screen IDs newline
        first_cell = fid + ("\n" + "\n".join(screens) if screens else "")
        is_skipped = include in ("미포함", "보류")

        for i, (label, content) in enumerate(items):
            row_vals = [
                first_cell if i == 0 else "",
                section if i == 0 else "",
                name if i == 0 else "",
                label,
                content,
                prio if i == 0 else "",
                api if i == 0 else "",
                include if i == 0 else "",
                status if i == 0 else "",
                memo if i == 0 else "",
            ]
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c_idx, value=val)
                cell.alignment = WRAP_TOP
                if is_skipped:
                    cell.font = GRAY_FONT
            r += 1

        # 보류/미포함은 빈 행 3개 추가 (인우님 패턴)
        if is_skipped:
            for _ in range(3):
                r += 1

    # ============= Sheet 3: DB구조 =============
    ws_db = wb.create_sheet("DB구조")
    ws_db.cell(row=1, column=1, value=f"{PROJECT_NAME} — DB 구조").font = TITLE_FONT
    ws_db.cell(row=2, column=1, value="상세는 docs/database.md 참고 (AWS RDS Postgres + Cognito + Lambda 권장)").font = GUIDE_FONT
    ws_db.cell(row=4, column=1, value="주요 테이블").font = HEADER_FONT
    db_tables = [
        "users / profiles", "folders", "folder_members", "folder_invites",
        "links", "ai_jobs"
    ]
    for i, name in enumerate(db_tables, 5):
        ws_db.cell(row=i, column=1, value=f"- {name}")
    ws_db.column_dimensions["A"].width = 80

    # ============= Sheet 4: 도메인 데이터셋 =============
    ws_dom = wb.create_sheet("도메인 데이터셋")
    ws_dom.cell(row=1, column=1, value=f"{PROJECT_NAME} — 도메인 데이터셋 (별도 작성)").font = TITLE_FONT
    ws_dom.cell(row=3, column=1, value="예시: AI provider 목록 / 강조색 팔레트 / 시드 데이터 정의 등").font = GUIDE_FONT
    ws_dom.column_dimensions["A"].width = 80

    # ============= Sheet 5: 기능명세 질의 =============
    ws_q = wb.create_sheet("기능명세 질의")
    headers_q = ["순번", "기능ID / (#화면ID)", "섹션", "기능명", "문의 제목", "문의사항", "답변"]
    for c, h in enumerate(headers_q, 1):
        style_header(ws_q.cell(row=1, column=c, value=h))
    widths_q = {1: 8, 2: 22, 3: 12, 4: 22, 5: 28, 6: 60, 7: 60}
    for col, w in widths_q.items():
        ws_q.column_dimensions[get_column_letter(col)].width = w
    ws_q.row_dimensions[1].height = 30
    for i in range(1, 31):
        ws_q.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
    ws_q.freeze_panes = "B2"

    # ============= Sheet 6: 변경 이력 =============
    ws_h = wb.create_sheet("변경 이력")
    headers_h = ["버전", "일자", "변경 내용", "변경자"]
    for c, h in enumerate(headers_h, 1):
        style_header(ws_h.cell(row=1, column=c, value=h))
    widths_h = {1: 10, 2: 14, 3: 80, 4: 14}
    for col, w in widths_h.items():
        ws_h.column_dimensions[get_column_letter(col)].width = w
    ws_h.row_dimensions[1].height = 28
    ws_h.cell(row=2, column=1, value=f"v{VERSION}")
    ws_h.cell(row=2, column=2, value=TODAY)
    ws_h.cell(row=2, column=3, value="최초 작성. 40개 기능. 모바일·웹 동시 PWA + 모바일↔웹 동기화 + 크롬 즐겨찾기 import 포함")
    ws_h.cell(row=2, column=4, value="인우")

    ws_h.cell(row=3, column=1, value=f"v{VERSION}.1")
    ws_h.cell(row=3, column=2, value=TODAY)
    ws_h.cell(row=3, column=3, value="결정 12건 반영: 이메일 6자리·게스트 출시 제거(Won't)·AI 폴백(MVP=mock+API키, v1.x=외부세션)·FCM Topic 분배+모바일/PC 채널 분리·비밀번호 변경 시 강제 로그아웃·즉시 hard delete·AWS API Gateway WebSocket·last-write-wins·공유멤버 50명·강조색 3색(핑크/블루/그레이)·단건 외부세션 미사용")
    ws_h.cell(row=3, column=4, value="인우")

    for r in (2, 3):
        for c in range(1, 5):
            ws_h.cell(row=r, column=c).alignment = WRAP_TOP

    wb.save(out_path)
    print(f"OK: {out_path}")
    print(f"   Sheet 1: 기능명세_목록 ({len(LIST_ROWS)}행)")
    detail_rows = sum(len(f[9]) + (3 if f[6] in ('미포함', '보류') else 0) for f in DETAIL_FEATURES)
    print(f"   Sheet 2: 기능명세_상세 (~{detail_rows}행)")
    print(f"   Sheet 3: DB구조")
    print(f"   Sheet 4: 도메인 데이터셋")
    print(f"   Sheet 5: 기능명세 질의 (1~30 미리 채움)")
    print(f"   Sheet 6: 변경 이력")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, "05_상세기능명세서.xlsx")
    build_workbook(out)
